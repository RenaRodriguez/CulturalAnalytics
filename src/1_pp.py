import openpyxl
import os
import csv

verzeichnis = "data/testdaten" 
dateien = [f for f in os.listdir(verzeichnis) if f.endswith('.xlsx')]

ergebnisse = []

# Hier mit openpyxl da wir die Informationen neben der Zelle "Herkunft" und "Datierung" brauchen sowie alle nicht leeren Zellen unter "Übersetzung". 
# Leider sind diese Zellen nicht immer in der gleichen Position, da die Zeile abhängig davon ist, wie viele Texte(varianten / funde) es gibt. 

for datei in dateien:
    dateipfad = os.path.join(verzeichnis, datei)
    print(datei)
    wb = openpyxl.load_workbook(dateipfad, data_only=True)
    
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        herkunft_list = []
        datierung_list = []
        uebersetzung = None

        for row_idx, row in enumerate(ws.iter_rows()):
            for i, cell in enumerate(row[:-1]):
                val = cell.value
                next_val = row[i+1].value if i + 1 < len(row) else None
                text = str(val).strip().lower() if isinstance(val, str) else ""
                next_text = str(next_val).strip().lower() if isinstance(next_val, str) else ""

                if "herkunft" in text and i + 1 < len(row):
                    eintrag = row[i + 1].value
                    if eintrag:
                        herkunft_list.append(str(eintrag).strip())

                if "datierung" in text and i + 1 < len(row):
                    eintrag = row[i + 1].value
                    if eintrag:
                        datierung_list.append(str(eintrag).strip())

                if text == "transliteration" and next_text == "übersetzung":
                    uebersetzungen = []
                    col = i + 2
                    row_pointer = row_idx + 2
                    while row_pointer <= ws.max_row:
                        val = ws.cell(row=row_pointer, column=col).value
                        if val is None or str(val).strip() == "":
                            break
                        uebersetzungen.append(str(val).strip())
                        row_pointer += 1
                    if uebersetzungen:
                        uebersetzung = ", ".join(uebersetzungen)

        # gleiche Länge, da jede Herkunft eine Datierung braucht und vv
        max_len = max(len(herkunft_list), len(datierung_list), 1)
        herkunft_list += [None] * (max_len - len(herkunft_list))
        datierung_list += [None] * (max_len - len(datierung_list))

        # Zusammensetzen
        for h, d in zip(herkunft_list, datierung_list):
            ergebnisse.append({
                "Datei": os.path.splitext(datei)[0],
                "Herkunft": h,
                "Datierung": d,
                "Übersetzung": uebersetzung
            })


# CSV
os.makedirs("data", exist_ok=True)
with open("data/extrahierte_testdaten.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=["Datei", "Herkunft", "Datierung", "Übersetzung"])
    writer.writeheader()
    for idx, eintrag in enumerate(ergebnisse, start=1):
        try:
            writer.writerow(eintrag)
            f.flush()
            print(f"Eintrag {idx}/{len(ergebnisse)} geschrieben: {eintrag}")
        except Exception as e:
            print(f"Fehler bei Eintrag {idx}: {eintrag}")
            print(e)
            break

print("Datei erstellt.")


